"""
BOHO CHIC — App de Análisis de Inventario y Traslados
======================================================
Desarrollado en el marco del Programa IA para la Productividad
Colsubsidio 2026 | Great Boost Inc.
"""

import streamlit as st
import pandas as pd
import json
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ── CONFIGURACIÓN DE PÁGINA ────────────────────────────────
st.set_page_config(
    page_title="BOHO CHIC — Inventario",
    page_icon="👗",
    layout="wide"
)

# ── ESTILOS ────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
        padding: 24px 32px; border-radius: 12px; margin-bottom: 24px;
        color: white;
    }
    .main-header h1 { font-size: 1.8rem; margin: 0; font-weight: 700; }
    .main-header p  { font-size: 0.9rem; opacity: 0.85; margin: 4px 0 0 0; }
    .btn-section {
        background: white; border-radius: 12px; padding: 24px;
        border: 1px solid #e0e6f0; margin-bottom: 20px;
    }
    .btn-title { font-size: 1.1rem; font-weight: 700; color: #1F3864; margin-bottom: 4px; }
    .btn-desc  { font-size: 0.85rem; color: #666; margin-bottom: 16px; }
    .metric-box {
        background: #f7f9fc; border-radius: 10px; padding: 16px;
        border: 1px solid #dde3ec; text-align: center;
    }
    .metric-val { font-size: 1.8rem; font-weight: 700; color: #1F3864; }
    .metric-lbl { font-size: 0.72rem; color: #888; text-transform: uppercase; letter-spacing: 0.05em; }
    .step-box {
        background: #f7f9fc; border-left: 4px solid #2E75B6;
        padding: 12px 16px; border-radius: 0 8px 8px 0; margin-bottom: 8px;
        font-size: 0.85rem; color: #333;
    }
</style>
""", unsafe_allow_html=True)

# ── HEADER ─────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>👗 BOHO CHIC — Análisis de Inventario</h1>
    <p>Programa IA para la Productividad · Colsubsidio 2026 · Great Boost Inc.</p>
</div>
""", unsafe_allow_html=True)

# ── CONSTANTES ─────────────────────────────────────────────
TALLAS = ['00-','99-UNICA','91-XS','92-S','93-M','94-L','95-XL',
          '04-04','06-06','06-6','08-08','08-8','10-10','12-12','14-14',
          '35-35','36-36','37-37','38-38','39-39','40-40','41-41']
N_TALLAS = len(TALLAS)

AZUL_OSC  = "1F3864"; AZUL_MED = "2E75B6"; VERDE_OSC = "375623"
AZUL_CLR  = "BDD7EE"; VERDE_CLR = "C6EFCE"; BLANCO = "FFFFFF"
GRIS_CLR  = "F2F2F2"; TOTAL_BG = "D9E1F2"; ALT_BG = "F5F8FF"; VERDE_LT = "E2EFDA"
NARANJA   = "ED7D31"

def tb():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, text, bg, fg=BLANCO, size=9):
    cell.value = text
    cell.font = Font(name='Arial', bold=True, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = tb()

def dc(cell, val, align='center', bold=False, bg=None, fg='000000', size=8):
    cell.value = val
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = tb()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

# ── FUNCIONES DE PROCESAMIENTO ─────────────────────────────

@st.cache_data(show_spinner=False)
def procesar_datos(stock_bytes, ventas_bytes, prio_bytes):
    """Carga y procesa los tres archivos de entrada."""

    prio_df = pd.read_excel(io.BytesIO(prio_bytes))
    prio_df['CONCA']       = prio_df['Conca'].astype(str).str.strip()
    prio_df['OBSERVACION'] = prio_df['Observación'].astype(str).str.strip()
    prio_map = dict(zip(prio_df['CONCA'], prio_df['OBSERVACION']))
    prio_set = set(prio_df['CONCA'])

    stock_raw = pd.read_excel(io.BytesIO(stock_bytes), header=[0,1])
    store_col_map = {}
    for col in stock_raw.columns:
        if 'Control de Stocks:' in str(col[0]):
            tienda = col[0].replace('Control de Stocks: ', '').strip()
            store_col_map[tienda] = col
    tiendas = list(store_col_map.keys())

    stock = pd.DataFrame()
    stock['Referencia'] = stock_raw[('Campos Relacionados','Referencia')].astype(str).str.strip()
    stock['Descripcion']= stock_raw[('Campos Relacionados','Descripción')].astype(str).str.strip()
    stock['Color']      = stock_raw[('Campos Relacionados','Color')].astype(str).str.strip()
    stock['Talla']      = stock_raw[('Campos Relacionados','Talla')].astype(str).str.strip()
    stock['Familia']    = stock_raw[('Campos Relacionados','Família')].astype(str).str.strip()
    stock['CONCA']      = stock['Referencia'] + '-' + stock['Color']
    for t in tiendas:
        stock[t] = pd.to_numeric(stock_raw[store_col_map[t]], errors='coerce').fillna(0)
    stock = stock[stock['Referencia'].notna() & ~stock['Referencia'].isin(['nan',''])]

    ventas_raw = pd.read_excel(io.BytesIO(ventas_bytes))
    ventas = ventas_raw[ventas_raw['Referencia'].notna()].copy()
    ventas['Ref_clean']   = ventas['Referencia'].apply(lambda x: str(int(float(x))))
    ventas['Color_clean'] = ventas['Color'].astype(str).str.strip()
    ventas['CONCA']       = ventas['Ref_clean'] + '-' + ventas['Color_clean']
    ventas['Uds']         = pd.to_numeric(ventas['Uds.'], errors='coerce').fillna(0)
    ventas_pos = ventas[ventas['Uds'] > 0]

    vta_tienda = ventas_pos.groupby(['CONCA','PUNTO_DE_VENTA'])['Uds'].sum().reset_index()
    vta_tienda.columns = ['CONCA','TIENDA','VTA_TDA']
    vta_total  = ventas_pos.groupby('CONCA')['Uds'].sum().reset_index()
    vta_total.columns  = ['CONCA','VTA_SKU']
    concas_con_ventas  = set(ventas_pos['CONCA'].unique())

    stock_prio  = stock[stock['CONCA'].isin(prio_set)].copy()
    # desc_map includes ALL stock CONCAs so dashboard top5 shows proper descriptions
    desc_map    = stock.groupby('CONCA')['Descripcion'].first().to_dict()
    familia_map = stock_prio.groupby('CONCA')['Familia'].first().to_dict()

    conca_stk = {}
    for conca, grp in stock_prio.groupby('CONCA'):
        conca_stk[conca] = {}
        for talla in TALLAS:
            rows = grp[grp['Talla'] == talla]
            if len(rows) == 0:
                continue
            for t in tiendas:
                val = int(rows[t].sum())
                if val > 0:
                    if t not in conca_stk[conca]:
                        conca_stk[conca][t] = {}
                    conca_stk[conca][t][talla] = val

    stock_all = stock.groupby('CONCA')[tiendas].sum()
    stock_all['STK_TOTAL'] = stock_all[tiendas].sum(axis=1)
    stock_all = stock_all[stock_all['STK_TOTAL'] > 0]
    total_concas_stk     = len(stock_all)
    concas_quietos_total = len(stock_all.index.difference(pd.Index(list(concas_con_ventas))))

    tienda_quiet = {}
    for t in tiendas:
        mask  = stock_all[t] > 0
        q_idx = stock_all[mask].index.difference(pd.Index(list(concas_con_ventas)))
        cnt   = len(q_idx); stk = int(stock_all.loc[q_idx, t].sum())
        if cnt > 0:
            top = stock_all.loc[q_idx, t].nlargest(5)
            top_list = [{"conca": c, "desc": desc_map.get(c, c), "stk": int(v)} for c,v in top.items()]
            tienda_quiet[t] = {"concas_quietos": cnt, "stk_parado": stk, "top_concas": top_list}

    conca_vta_map  = vta_total.set_index('CONCA')['VTA_SKU'].to_dict()
    vta_tienda_map = vta_tienda.groupby('CONCA').apply(
        lambda x: dict(zip(x['TIENDA'], x['VTA_TDA']))).to_dict()
    prio_concas    = sorted(prio_set, key=lambda c: -conca_vta_map.get(c, 0))

    all_blocks = []
    for conca in prio_concas:
        obs     = prio_map.get(conca, '')
        desc    = desc_map.get(conca, conca)
        familia = familia_map.get(conca, '')
        vta_sku = int(conca_vta_map.get(conca, 0))
        t_vtas  = vta_tienda_map.get(conca, {})
        tiendas_stk   = set(conca_stk.get(conca, {}).keys())
        todas_tiendas = tiendas_stk | set(t_vtas.keys())
        if not todas_tiendas:
            continue
        s_tdas = len(tiendas_stk)
        tiendas_sorted = sorted(todas_tiendas, key=lambda t: -t_vtas.get(t, 0))

        block = []
        for i, t in enumerate(tiendas_sorted):
            talla_stk = conca_stk.get(conca, {}).get(t, {})
            block.append({'CONCA': conca, 'SKU': 1 if i==0 else 0,
                'CONCA_COD_TDA': f"{conca}-{t}", 'MODELO_COLOR': desc,
                'COD_TIENDA': t, 'NOMBRE_TIENDA': t, 'COLECCION': familia,
                'VTA_TDA': int(t_vtas.get(t, 0)), 'VTA_SKU': vta_sku,
                'S_TDAS': s_tdas, 'OBSERVACION': obs,
                'tallas': {tl: talla_stk.get(tl, 0) for tl in TALLAS},
                'is_total': False})

        total_tallas = {tl: sum(conca_stk.get(conca,{}).get(t,{}).get(tl,0)
                                for t in todas_tiendas) for tl in TALLAS}
        block.append({'CONCA': conca, 'SKU': 0,
            'CONCA_COD_TDA': f"{conca}-TOTAL", 'MODELO_COLOR': desc,
            'COD_TIENDA': 'TOTAL', 'NOMBRE_TIENDA': 'TOTAL', 'COLECCION': familia,
            'VTA_TDA': None, 'VTA_SKU': vta_sku, 'S_TDAS': s_tdas,
            'OBSERVACION': obs, 'tallas': total_tallas, 'is_total': True})
        all_blocks.append(block)

    return {
        'all_blocks': all_blocks,
        'tienda_quiet': tienda_quiet,
        'total_concas_stk': total_concas_stk,
        'concas_quietos_total': concas_quietos_total,
        'prio_set': prio_set,
        'concas_con_ventas': concas_con_ventas,
        'tiendas': tiendas,
    }


def generar_excel(data):
    """Genera el Excel con ANALISIS TRS, RESUMEN CONCAs y RESUMEN TIENDA."""
    wb = Workbook()
    all_blocks         = data['all_blocks']
    tienda_quiet       = data['tienda_quiet']
    total_concas_stk   = data['total_concas_stk']
    prio_set           = data['prio_set']
    concas_con_ventas  = data['concas_con_ventas']

    COL_INI_START = 13; COL_INI_END = 34
    COL_FIN_START = 35; COL_FIN_END = 56

    # ── HOJA 1: ANALISIS TRS ──
    ws = wb.active
    ws.title = "ANALISIS TRS"
    ws.freeze_panes = "A3"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    hdr(ws.cell(1,1), "INFORMACIÓN", AZUL_OSC, size=10)
    ws.merge_cells(start_row=1, start_column=COL_INI_START, end_row=1, end_column=COL_INI_END)
    hdr(ws.cell(1,COL_INI_START), "STK INICIAL", AZUL_MED, size=10)
    ws.merge_cells(start_row=1, start_column=COL_FIN_START, end_row=1, end_column=COL_FIN_END)
    hdr(ws.cell(1,COL_FIN_START), "STK FINAL  (editable — ingrese traslados aquí)", VERDE_OSC, size=10)
    ws.row_dimensions[1].height = 22

    INFO_H = ['MARCA','CONCA','SKU','CONCA - COD. TDA','MODELO-COLOR','COD. TIENDA',
              'NOMBRE TIENDA','COLECCIÓN','VTA TDA','VTA SKU','S. TDAS','OBSERVACIÓN']
    INFO_W = [12,20,5,28,32,16,22,14,9,9,8,18]
    for i,(h,w) in enumerate(zip(INFO_H,INFO_W),1):
        hdr(ws.cell(2,i), h, AZUL_OSC, size=8)
        ws.column_dimensions[get_column_letter(i)].width = w
    for i,tl in enumerate(TALLAS, COL_INI_START):
        hdr(ws.cell(2,i), tl, AZUL_CLR, fg=AZUL_OSC, size=8)
        ws.column_dimensions[get_column_letter(i)].width = 6
    for i,tl in enumerate(TALLAS, COL_FIN_START):
        hdr(ws.cell(2,i), tl, VERDE_CLR, fg=VERDE_OSC, size=8)
        ws.column_dimensions[get_column_letter(i)].width = 6
    ws.row_dimensions[2].height = 22

    current_row = 3
    boundaries  = []

    for bidx, block in enumerate(all_blocks):
        b_start = current_row
        tot_row = None
        alt     = bidx % 2 != 0

        for rd in block:
            r      = current_row
            is_tot = rd['is_total']
            rbg    = TOTAL_BG if is_tot else (ALT_BG if alt else None)
            fbold  = is_tot
            fcol   = AZUL_OSC if is_tot else '000000'

            info_vals  = ['FREE SPIRIT', rd['CONCA'], rd['SKU'], rd['CONCA_COD_TDA'],
                          rd['MODELO_COLOR'], rd['COD_TIENDA'], rd['NOMBRE_TIENDA'],
                          rd['COLECCION'], rd['VTA_TDA'], rd['VTA_SKU'],
                          rd['S_TDAS'], rd['OBSERVACION']]
            info_aligns= ['center','left','center','center','left','center',
                          'left','center','center','center','center','left']
            for ci,(v,al) in enumerate(zip(info_vals,info_aligns),1):
                dc(ws.cell(r,ci), v, align=al, bold=fbold, bg=rbg, fg=fcol)

            for ci,tl in enumerate(TALLAS, COL_INI_START):
                v = rd['tallas'].get(tl,0)
                dc(ws.cell(r,ci), int(v) if v else '', bg=rbg, fg=fcol, bold=fbold)

            if not is_tot:
                for ci,tl in enumerate(TALLAS, COL_FIN_START):
                    v = rd['tallas'].get(tl,0)
                    dc(ws.cell(r,ci), int(v) if v else '', bg=VERDE_LT if alt else None)
            else:
                tot_row = r
                for ci in range(COL_FIN_START, COL_FIN_END+1):
                    c = ws.cell(r,ci)
                    c.font = Font(name='Arial', size=8, bold=True, color=AZUL_OSC)
                    c.fill = PatternFill("solid", fgColor=TOTAL_BG)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = tb()

            ws.row_dimensions[r].height = 14
            current_row += 1

        boundaries.append((b_start, tot_row))

    for b_start, tot_r in boundaries:
        if tot_r is None:
            continue
        for ci_off in range(N_TALLAS):
            fin_col = COL_FIN_START + ci_off
            col_l   = get_column_letter(fin_col)
            ws.cell(tot_r, fin_col).value = f"=SUM({col_l}{b_start}:{col_l}{tot_r-1})"

    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    RED_FONT = Font(name='Arial', size=8, bold=True, color="FFFFFF")
    max_row  = current_row - 1
    for ci_off in range(N_TALLAS):
        ini_l = get_column_letter(COL_INI_START + ci_off)
        fin_l = get_column_letter(COL_FIN_START + ci_off)
        ws.conditional_formatting.add(
            f"{fin_l}3:{fin_l}{max_row}",
            FormulaRule(formula=[f'AND($F3="TOTAL",{fin_l}3<>{ini_l}3)'],
                        fill=RED_FILL, font=RED_FONT))

    # ── HOJA 2: RESUMEN CONCAs ──
    ws2 = wb.create_sheet("RESUMEN CONCAs")
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:F1")
    hdr(ws2.cell(1,1), f"RESUMEN — {len(all_blocks)} CONCAs priorizados", AZUL_OSC, size=11)
    ws2.row_dimensions[1].height = 24
    h2=['CONCA','DESCRIPCIÓN','OBSERVACIÓN','VTA SKU (4 sem)','# TIENDAS CON STK','QUIETO']
    w2=[22,38,20,16,18,10]
    for i,(h,w) in enumerate(zip(h2,w2),1):
        hdr(ws2.cell(2,i), h, AZUL_MED, size=9)
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 20
    for ri,block in enumerate(all_blocks,3):
        rd=block[0]; conca=rd['CONCA']
        quieto='SÍ' if conca not in concas_con_ventas else 'NO'
        alt=ri%2==0; bg=GRIS_CLR if alt else None
        vals=[conca,rd['MODELO_COLOR'],rd['OBSERVACION'],rd['VTA_SKU'],rd['S_TDAS'],quieto]
        als=['center','left','center','center','center','center']
        for ci,(v,al) in enumerate(zip(vals,als),1):
            c=ws2.cell(ri,ci); dc(c,v,align=al,bg=bg,size=9)
            if ci==6 and quieto=='SÍ':
                c.fill=PatternFill("solid",fgColor="FCE4D6")
                c.font=Font(name='Arial',size=9,bold=True,color="C00000")
        ws2.row_dimensions[ri].height=15

    # ── HOJA 3: RESUMEN TIENDA ──
    ws3 = wb.create_sheet("RESUMEN TIENDA")
    ws3.freeze_panes = "A3"
    ws3.merge_cells("A1:E1")
    hdr(ws3.cell(1,1), "CONCAs QUIETOS POR TIENDA — Últimas 4 semanas", AZUL_OSC, size=11)
    ws3.row_dimensions[1].height = 24
    h3=['TIENDA','CONCAs QUIETOS','STOCK PARADO (uds)',
        '% del inventario quieto total','PRIORIDAD']
    w3=[24,18,20,28,14]
    for i,(h,w) in enumerate(zip(h3,w3),1):
        hdr(ws3.cell(2,i),h,AZUL_MED,size=9)
        ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.row_dimensions[2].height=20
    tienda_list=sorted(tienda_quiet.items(),key=lambda x:-x[1]['concas_quietos'])
    total_quietos = sum(d['concas_quietos'] for d in tienda_quiet.values())
    for ri,(tn,d) in enumerate(tienda_list,3):
        pct=d['concas_quietos']/total_quietos if total_quietos > 0 else 0
        pri="ALTA" if pct>0.20 else ("MEDIA" if pct>0.08 else "BAJA")
        alt=ri%2==0; bg=GRIS_CLR if alt else None
        vals=[tn,d['concas_quietos'],d['stk_parado'],pct,pri]
        als=['left','center','center','center','center']
        for ci,(v,al) in enumerate(zip(vals,als),1):
            c=ws3.cell(ri,ci); dc(c,v,align=al,bg=bg,size=9)
            if ci==4: c.number_format='0.0%'
        ws3.row_dimensions[ri].height=16
    tr=len(tienda_list)+3
    for ci,v in enumerate(["TOTAL",
                            sum(d['concas_quietos'] for d in tienda_quiet.values()),
                            sum(d['stk_parado'] for d in tienda_quiet.values()),"",""],1):
        c=ws3.cell(tr,ci,v)
        c.font=Font(name='Arial',bold=True,size=9,color=AZUL_OSC)
        c.fill=PatternFill("solid",fgColor=AZUL_CLR); c.border=tb()
        c.alignment=Alignment(horizontal='center',vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_html(data):
    """Genera el dashboard HTML interactivo."""
    tienda_quiet       = data['tienda_quiet']
    total_concas_stk   = data['total_concas_stk']
    concas_quietos_total = data['concas_quietos_total']
    prio_set           = data['prio_set']

    bubble_data = [{"tienda":t,"concas_quietos":d["concas_quietos"],
                    "stk_parado":d["stk_parado"],"top_concas":d["top_concas"]}
                   for t,d in tienda_quiet.items()]
    bj = json.dumps(bubble_data, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>BOHO CHIC — Dashboard CONCAs Quietos</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f0f2f5;color:#1a1a2e}}
header{{background:linear-gradient(135deg,#1F3864,#2E75B6);color:white;padding:18px 32px;
  display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 8px rgba(0,0,0,.2)}}
header h1{{font-size:1.3rem;font-weight:700}}
header p{{font-size:.8rem;opacity:.85;margin-top:2px}}
.badge{{background:#ED7D31;color:white;border-radius:20px;padding:6px 16px;font-size:.85rem;font-weight:700}}
.kpi-bar{{display:flex;gap:16px;padding:16px 32px;background:white;border-bottom:1px solid #dde3ec;flex-wrap:wrap}}
.kpi{{text-align:center;padding:10px 24px;border-radius:10px;background:#f7f9fc;border:1px solid #dde3ec;min-width:130px}}
.kpi .val{{font-size:1.6rem;font-weight:700;color:#1F3864}}
.kpi .lbl{{font-size:.7rem;color:#666;text-transform:uppercase;letter-spacing:.05em;margin-top:2px}}
.main{{display:flex;height:calc(100vh - 140px)}}
#bubble-panel{{flex:1;padding:24px 32px;overflow-y:auto;position:relative}}
#bubble-panel h2{{font-size:.9rem;color:#555;margin-bottom:16px;font-weight:600;text-transform:uppercase;letter-spacing:.04em}}
#bubble-canvas{{position:relative;width:100%;min-height:520px}}
.bubble{{position:absolute;border-radius:50%;cursor:pointer;display:flex;flex-direction:column;
  align-items:center;justify-content:center;transition:transform .2s,box-shadow .2s;
  border:3px solid rgba(255,255,255,.6);text-align:center;overflow:hidden;user-select:none}}
.bubble:hover{{transform:scale(1.08);box-shadow:0 6px 24px rgba(0,0,0,.25);z-index:10}}
.bubble.selected{{border:3px solid #1F3864;box-shadow:0 0 0 4px rgba(31,56,100,.25);z-index:11}}
.bubble .b-count{{font-weight:700;color:white;line-height:1.1}}
.bubble .b-name{{color:rgba(255,255,255,.92);font-weight:600;line-height:1.2;text-transform:uppercase;letter-spacing:.02em;padding:0 4px}}
#detail-panel{{width:360px;background:white;border-left:1px solid #dde3ec;display:flex;flex-direction:column;overflow:hidden;flex-shrink:0}}
#detail-header{{background:#1F3864;color:white;padding:16px 20px}}
#detail-header h3{{font-size:1rem;font-weight:700}}
#detail-header p{{font-size:.75rem;opacity:.8;margin-top:3px}}
#detail-body{{flex:1;overflow-y:auto;padding:16px}}
.stat-row{{display:flex;gap:10px;margin-bottom:16px}}
.stat-box{{flex:1;background:#f7f9fc;border-radius:8px;padding:10px;text-align:center;border:1px solid #e0e6f0}}
.stat-box .sv{{font-size:1.4rem;font-weight:700;color:#1F3864}}
.stat-box .sl{{font-size:.68rem;color:#888;text-transform:uppercase;letter-spacing:.04em}}
#detail-body h4{{font-size:.78rem;text-transform:uppercase;letter-spacing:.05em;color:#888;margin-bottom:8px;font-weight:700}}
.conca-row{{display:flex;align-items:center;gap:10px;padding:9px 10px;border-radius:7px;background:#f7f9fc;margin-bottom:6px;border:1px solid #e8edf5}}
.conca-num{{background:#ED7D31;color:white;border-radius:50%;width:22px;height:22px;display:flex;align-items:center;justify-content:center;font-size:.7rem;font-weight:700;flex-shrink:0}}
.conca-info{{flex:1;min-width:0}}
.conca-info .cn{{font-size:.72rem;font-weight:700;color:#1F3864;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.conca-info .cd{{font-size:.68rem;color:#666;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.conca-stk{{font-size:.78rem;font-weight:700;color:#ED7D31;flex-shrink:0}}
.empty-state{{display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;color:#bbb;text-align:center;padding:32px}}
.empty-state p{{font-size:.85rem;line-height:1.5}}
.priority{{display:inline-block;border-radius:12px;padding:3px 10px;font-size:.7rem;font-weight:700;margin-top:6px}}
.p-high{{background:#fde8e8;color:#c0392b}}.p-med{{background:#fef3cd;color:#8a6d00}}.p-low{{background:#e8f5e9;color:#2e7d32}}
.legend{{position:absolute;bottom:12px;left:12px;background:rgba(255,255,255,.92);border-radius:8px;padding:10px 14px;font-size:.72rem;color:#555;border:1px solid #dde3ec}}
.legend-title{{font-weight:700;color:#1F3864;margin-bottom:6px}}
.legend-row{{display:flex;align-items:center;gap:7px;margin-bottom:4px}}
.legend-dot{{border-radius:50%;flex-shrink:0}}
</style></head><body>
<header>
  <div><h1>BOHO CHIC — CONCAs sin movimiento en tiendas</h1>
  <p>Haga clic en una tienda para ver el detalle de CONCAs quietos</p></div>
  <div class="badge">{concas_quietos_total:,} CONCAs quietos</div>
</header>
<div class="kpi-bar">
  <div class="kpi"><div class="val">{concas_quietos_total:,}</div><div class="lbl">CONCAs quietos</div></div>
  <div class="kpi"><div class="val">{total_concas_stk:,}</div><div class="lbl">CONCAs con stock</div></div>
  <div class="kpi"><div class="val">{round(concas_quietos_total/total_concas_stk*100)}%</div><div class="lbl">% quietos / total</div></div>
  <div class="kpi"><div class="val">{len(tienda_quiet)}</div><div class="lbl">Tiendas con quietos</div></div>
  <div class="kpi"><div class="val">{len(prio_set)}</div><div class="lbl">CONCAs priorizados</div></div>
</div>
<div class="main">
  <div id="bubble-panel">
    <h2>Tamaño = Nº de CONCAs quietos · Haga clic en una tienda para ver detalle</h2>
    <div id="bubble-canvas"></div>
    <div class="legend">
      <div class="legend-title">Prioridad</div>
      <div class="legend-row"><div class="legend-dot" style="width:12px;height:12px;background:#c0392b"></div> ALTA &gt;20% del inventario quieto</div>
      <div class="legend-row"><div class="legend-dot" style="width:12px;height:12px;background:#e67e22"></div> MEDIA 8%-20%</div>
      <div class="legend-row"><div class="legend-dot" style="width:12px;height:12px;background:#27ae60"></div> BAJA &lt;8%</div>
    </div>
  </div>
  <div id="detail-panel">
    <div id="detail-header"><h3 id="dh-title">Detalle de tienda</h3><p id="dh-sub">Seleccione una tienda</p></div>
    <div id="detail-body"><div class="empty-state"><p>Haga clic en cualquier<br>burbuja para ver los<br>CONCAs quietos</p></div></div>
  </div>
</div>
<script>
const DATA={bj};
const TOTAL_QUIETOS={concas_quietos_total};
const maxVal=Math.max(...DATA.map(d=>d.concas_quietos));
function getColor(q){{const p=q/TOTAL_QUIETOS;if(p>.20)return{{bg:'linear-gradient(135deg,#c0392b,#e74c3c)',border:'#922b21'}};if(p>.08)return{{bg:'linear-gradient(135deg,#e67e22,#f39c12)',border:'#b9770e'}};return{{bg:'linear-gradient(135deg,#27ae60,#2ecc71)',border:'#1e8449'}}}}
function getPriority(q){{const p=q/TOTAL_QUIETOS;if(p>.20)return'<span class="priority p-high">🔴 ALTA</span>';if(p>.08)return'<span class="priority p-med">🟡 MEDIA</span>';return'<span class="priority p-low">🟢 BAJA</span>'}}
function getSize(q){{return 52+Math.sqrt(q/maxVal)*98}}
function layoutBubbles(){{
  const canvas=document.getElementById('bubble-canvas');
  const W=canvas.offsetWidth;
  const items=[...DATA].sort((a,b)=>b.concas_quietos-a.concas_quietos);
  const placed=[];let maxY=0;
  items.forEach((d,i)=>{{
    const r=getSize(d.concas_quietos)/2;let ok=false;
    for(let a=0;a<2000&&!ok;a++){{
      const x=r+Math.random()*(W-2*r),y=r+Math.random()*480;
      if(!placed.some(p=>{{const dx=x-p.x,dy=y-p.y;return Math.sqrt(dx*dx+dy*dy)<r+p.r+8}}))
        {{placed.push({{x,y,r,data:d}});if(y+r>maxY)maxY=y+r;ok=true}}
    }}
    if(!ok){{const x=r+(i%8)*(W/8),y=r+Math.floor(i/8)*160+maxY;placed.push({{x,y,r,data:d}});if(y+r>maxY)maxY=y+r}}
  }});
  canvas.style.height=(maxY+20)+'px';
  placed.forEach(({{x,y,r,data:d}})=>{{
    const color=getColor(d.concas_quietos),size=r*2;
    const div=document.createElement('div');div.className='bubble';
    div.style.cssText=`left:${{x-r}}px;top:${{y-r}}px;width:${{size}}px;height:${{size}}px;background:${{color.bg}};border-color:${{color.border}};`;
    const fs=Math.max(8,Math.min(13,size/7)),cs=Math.max(11,Math.min(22,size/5));
    div.innerHTML=`<span class="b-count" style="font-size:${{cs}}px">${{d.concas_quietos}}</span><span class="b-name" style="font-size:${{fs}}px">${{d.tienda}}</span>`;
    div.addEventListener('click',()=>showDetail(d,div));canvas.appendChild(div);
  }});
}}
function showDetail(d,el){{
  document.querySelectorAll('.bubble').forEach(b=>b.classList.remove('selected'));el.classList.add('selected');
  document.getElementById('dh-title').textContent=d.tienda;
  document.getElementById('dh-sub').textContent=d.concas_quietos+' CONCAs sin venta en 4 semanas';
  const pct=((d.concas_quietos/TOTAL_QUIETOS)*100).toFixed(1);
  document.getElementById('detail-body').innerHTML=`
    <div class="stat-row">
      <div class="stat-box"><div class="sv">${{d.concas_quietos}}</div><div class="sl">CONCAs quietos</div></div>
      <div class="stat-box"><div class="sv">${{d.stk_parado.toLocaleString()}}</div><div class="sl">Uds. paradas</div></div>
      <div class="stat-box"><div class="sv">${{pct}}%</div><div class="sl">% del inventario quieto</div></div>
    </div>
    ${{getPriority(d.concas_quietos)}}
    <br><h4 style="margin-top:12px">Top 5 CONCAs con más stock parado</h4>
    ${{d.top_concas.map((c,i)=>`
      <div class="conca-row">
        <div class="conca-num">${{i+1}}</div>
        <div class="conca-info"><div class="cn">${{c.conca}}</div><div class="cd">${{c.desc}}</div></div>
        <div class="conca-stk">${{c.stk}} uds</div>
      </div>`).join('')}}
    <br><p style="font-size:.7rem;color:#999;line-height:1.5">💡 Consulte la hoja <strong>ANALISIS TRS</strong> del Excel para registrar traslados.</p>`;
}}
window.addEventListener('load',()=>{{layoutBubbles();window.addEventListener('resize',()=>{{document.getElementById('bubble-canvas').innerHTML='';document.getElementById('bubble-canvas').style.height='';layoutBubbles()}});}});
</script></body></html>"""
    return html.encode('utf-8')


def generar_registro_trs(excel_editado_bytes):
    """Lee el Excel editado, detecta diferencias STK FINAL vs STK INICIAL
    y genera la hoja REGISTRO TRS."""
    wb = load_workbook(io.BytesIO(excel_editado_bytes), data_only=True)
    ws = wb['ANALISIS TRS']

    # Leer encabezados fila 2 para encontrar columnas
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]

    def find_col(name):
        for i, v in enumerate(row2, 1):
            if v == name:
                return i
        return None

    col_conca      = find_col('CONCA')
    col_marca      = find_col('MARCA') or 1
    col_modelo     = find_col('MODELO-COLOR')
    col_tienda     = find_col('NOMBRE TIENDA')
    col_cod_tienda = find_col('COD. TIENDA')

    # Encontrar secciones STK INICIAL y STK FINAL por fila 1
    ini_start = fin_start = None
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    for i, v in enumerate(row1, 1):
        if v == 'STK INICIAL' and ini_start is None:
            ini_start = i
        if v == 'STK FINAL  (editable — ingrese traslados aquí)' and fin_start is None:
            fin_start = i

    if not ini_start or not fin_start:
        return None, "No se encontraron las secciones STK INICIAL / STK FINAL en el archivo."

    ini_end = ini_start + N_TALLAS - 1
    fin_end = fin_start + N_TALLAS - 1

    # Leer tallas de fila 2
    tallas_ini = [ws.cell(2, c).value for c in range(ini_start, ini_end+1)]

    # Leer datos fila por fila
    movimientos = []
    for r in range(3, ws.max_row+1):
        cod_tienda = ws.cell(r, col_cod_tienda).value if col_cod_tienda else None
        if cod_tienda == 'TOTAL' or cod_tienda is None:
            continue

        conca      = ws.cell(r, col_conca).value if col_conca else ''
        marca      = ws.cell(r, col_marca).value or 'FREE SPIRIT'
        modelo     = ws.cell(r, col_modelo).value if col_modelo else ''
        tienda     = ws.cell(r, col_tienda).value if col_tienda else cod_tienda

        for ci_off, talla in enumerate(tallas_ini):
            ini_val = ws.cell(r, ini_start + ci_off).value or 0
            fin_val = ws.cell(r, fin_start + ci_off).value or 0
            try:
                ini_val = int(ini_val); fin_val = int(fin_val)
            except:
                continue
            diff = fin_val - ini_val
            if diff != 0:
                movimientos.append({
                    'MARCA': marca, 'CONCA': conca, 'MODELO COLOR': modelo,
                    'TALLA': talla, 'TIENDA': tienda, 'DIFERENCIA': diff
                })

    if not movimientos:
        return None, "No se detectaron diferencias entre STK INICIAL y STK FINAL."

    # Emparejar orígenes y destinos por CONCA+TALLA
    import collections
    registro = []

    # Agrupar por CONCA + TALLA
    grupos = collections.defaultdict(lambda: {'ori': [], 'dst': []})
    for m in movimientos:
        key = (m['CONCA'], m['TALLA'])
        if m['DIFERENCIA'] < 0:
            grupos[key]['ori'].append({'tienda': m['TIENDA'], 'cant': abs(m['DIFERENCIA']),
                                       'marca': m['MARCA'], 'modelo': m['MODELO COLOR']})
        else:
            grupos[key]['dst'].append({'tienda': m['TIENDA'], 'cant': m['DIFERENCIA'],
                                       'marca': m['MARCA'], 'modelo': m['MODELO COLOR']})

    for (conca, talla), g in grupos.items():
        oris = list(g['ori'])
        dsts = list(g['dst'])
        if not oris or not dsts:
            continue

        marca  = oris[0]['marca']
        modelo = oris[0]['modelo']

        # Emparejar secuencialmente
        oi = di = 0
        cant_ori = oris[0]['cant'] if oris else 0
        cant_dst = dsts[0]['cant'] if dsts else 0

        while oi < len(oris) and di < len(dsts):
            mov = min(cant_ori, cant_dst)
            if mov > 0:
                registro.append({
                    'MARCA': marca, 'CONCA': conca, 'MODELO COLOR': modelo,
                    'TALLA': talla,
                    'ORIGEN': oris[oi]['tienda'],
                    'DESTINO': dsts[di]['tienda'],
                    'CANTIDAD': mov
                })
            cant_ori -= mov
            cant_dst -= mov
            if cant_ori == 0:
                oi += 1
                if oi < len(oris):
                    cant_ori = oris[oi]['cant']
            if cant_dst == 0:
                di += 1
                if di < len(dsts):
                    cant_dst = dsts[di]['cant']

    # Agregar hoja REGISTRO TRS al workbook
    if 'REGISTRO TRS' in wb.sheetnames:
        del wb['REGISTRO TRS']
    ws_reg = wb.create_sheet("REGISTRO TRS")

    ws_reg.merge_cells("A1:G1")
    hdr(ws_reg.cell(1,1), f"REGISTRO DE TRASLADOS — {len(registro)} movimientos detectados",
        NARANJA, size=11)
    ws_reg.row_dimensions[1].height = 26

    reg_h = ['MARCA','CONCA','MODELO COLOR','TALLA','ORIGEN','DESTINO','CANTIDAD']
    reg_w = [14, 22, 36, 10, 22, 22, 12]
    for i,(h,w) in enumerate(zip(reg_h,reg_w),1):
        hdr(ws_reg.cell(2,i), h, AZUL_OSC, size=9)
        ws_reg.column_dimensions[get_column_letter(i)].width = w
    ws_reg.row_dimensions[2].height = 20

    for ri, mov in enumerate(registro, 3):
        alt = ri % 2 == 0
        bg  = GRIS_CLR if alt else None
        vals = [mov['MARCA'], mov['CONCA'], mov['MODELO COLOR'],
                mov['TALLA'], mov['ORIGEN'], mov['DESTINO'], mov['CANTIDAD']]
        aligns = ['center','center','left','center','center','center','center']
        for ci,(v,al) in enumerate(zip(vals,aligns),1):
            dc(ws_reg.cell(ri,ci), v, align=al, bg=bg, size=9)
        ws_reg.row_dimensions[ri].height = 15

    # Mover REGISTRO TRS al frente
    wb.move_sheet("REGISTRO TRS", offset=-len(wb.sheetnames)+1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f"✅ {len(registro)} movimientos generados correctamente."


# ══════════════════════════════════════════════════════════
# INTERFAZ STREAMLIT
# ══════════════════════════════════════════════════════════

tab1, tab2 = st.tabs(["📊 Generar archivos de trabajo", "📋 Generar Registro de Traslados"])

# ── TAB 1: GENERAR ARCHIVOS ────────────────────────────────
with tab1:
    st.markdown("""
    <div class="step-box">📁 <b>Paso 1:</b> Sube los tres archivos exportados de ICG</div>
    <div class="step-box">⚙️ <b>Paso 2:</b> Haz clic en "Generar"</div>
    <div class="step-box">⬇️ <b>Paso 3:</b> Descarga el Excel de trabajo y el Dashboard HTML</div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        f_stock  = st.file_uploader("📦 Archivo de Stock", type=['xlsx','xls'], key='stock')
    with col2:
        f_ventas = st.file_uploader("💰 Archivo de Ventas (4 semanas)", type=['xlsx','xls'], key='ventas')
    with col3:
        f_prio   = st.file_uploader("🎯 Archivo de CONCAs Priorizados", type=['xlsx','xls'], key='prio')

    st.markdown("---")

    if f_stock and f_ventas and f_prio:
        if st.button("⚙️ Generar archivos de trabajo", type="primary", use_container_width=True):
            with st.spinner("Procesando datos..."):
                try:
                    data = procesar_datos(f_stock.read(), f_ventas.read(), f_prio.read())
                    st.session_state['data']        = data
                    st.session_state['excel_bytes'] = generar_excel(data)
                    st.session_state['html_bytes']  = generar_html(data)
                    st.session_state['generado']    = True
                except Exception as e:
                    st.error(f"Error al procesar los archivos: {e}")

    if st.session_state.get('generado'):
        data = st.session_state['data']
        col_a, col_b, col_c, col_d = st.columns(4)
        with col_a:
            st.markdown(f'<div class="metric-box"><div class="metric-val">{data["concas_quietos_total"]:,}</div><div class="metric-lbl">CONCAs quietos</div></div>', unsafe_allow_html=True)
        with col_b:
            st.markdown(f'<div class="metric-box"><div class="metric-val">{data["total_concas_stk"]:,}</div><div class="metric-lbl">CONCAs con stock</div></div>', unsafe_allow_html=True)
        with col_c:
            st.markdown(f'<div class="metric-box"><div class="metric-val">{len(data["all_blocks"]):,}</div><div class="metric-lbl">CONCAs priorizados</div></div>', unsafe_allow_html=True)
        with col_d:
            st.markdown(f'<div class="metric-box"><div class="metric-val">{len(data["tienda_quiet"])}</div><div class="metric-lbl">Tiendas con quietos</div></div>', unsafe_allow_html=True)

        st.markdown("---")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="⬇️ Descargar Excel de trabajo",
                data=st.session_state['excel_bytes'],
                file_name="BOHO_CHIC_Analisis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_dl2:
            st.download_button(
                label="⬇️ Descargar Dashboard HTML",
                data=st.session_state['html_bytes'],
                file_name="BOHO_CHIC_Dashboard.html",
                mime="text/html",
                use_container_width=True
            )
        st.success("✅ Archivos listos. Puedes descargar ambos independientemente.")
    elif not (f_stock and f_ventas and f_prio):
        st.info("👆 Sube los tres archivos para continuar.")

# ── TAB 2: REGISTRO DE TRASLADOS ──────────────────────────
with tab2:
    st.markdown("""
    <div class="step-box">✏️ <b>Paso 1:</b> Descarga el Excel de trabajo (Tab anterior) y modifica la columna <b>STK FINAL</b></div>
    <div class="step-box">📤 <b>Paso 2:</b> Sube aquí el Excel con tus cambios</div>
    <div class="step-box">📋 <b>Paso 3:</b> Descarga el Excel con la hoja REGISTRO TRS generada automáticamente</div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    f_editado = st.file_uploader(
        "📤 Sube el Excel con STK FINAL modificado",
        type=['xlsx','xls'], key='editado'
    )
    st.markdown("---")

    if f_editado:
        if st.button("📋 Generar Registro de Traslados", type="primary", use_container_width=True):
            with st.spinner("Detectando movimientos y generando registro..."):
                try:
                    result_bytes, mensaje = generar_registro_trs(f_editado.read())
                    if result_bytes:
                        st.success(mensaje)
                        st.download_button(
                            label="⬇️ Descargar Excel con Registro TRS",
                            data=result_bytes,
                            file_name="BOHO_CHIC_Con_Registro_TRS.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.warning(mensaje)
                except Exception as e:
                    st.error(f"Error al procesar el archivo: {e}")
    else:
        st.info("👆 Sube el Excel editado para continuar.")
